#!/usr/bin/env python3
"""
Sistema INTELIGENTE para carregar tabelas FATO no DuckDB.
Aprende e lembra dos mapeamentos, detecta padrões automaticamente.
Gera documentação, histórico e amostras em Excel.
"""
import pandas as pd
import duckdb
from pathlib import Path
from datetime import datetime
import logging
import re
import json
import hashlib
import sys
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class CarregadorInteligenteTabelas:
    """Carregador inteligente para tabelas fato com aprendizado automático"""
    
    # Mapeamento de abreviações de meses
    MESES_ABREV = {
        'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
        'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
    }
    
    MESES_NOME = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }
    
    def __init__(self):
        # Definir caminhos base
        script_dir = Path(__file__).parent
        self.base_path = script_dir.parent / "dados_brutos"
        
        # Estrutura de pastas
        self.db_path = self.base_path / "duckdb" / "database.duckdb"
        self.fato_path = self.base_path / "fato"
        self.metadata_path = self.base_path / "metadata"
        self.amostras_path = self.base_path / "amostras"
        
        # Criar pastas se não existirem
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self.fato_path.mkdir(parents=True, exist_ok=True)
        self.metadata_path.mkdir(parents=True, exist_ok=True)
        self.amostras_path.mkdir(parents=True, exist_ok=True)
        
        # Criar subpastas para cada tipo de fato
        self.despesa_saldo_path = self.fato_path / "despesa_saldo"
        self.receita_saldo_path = self.fato_path / "receita_saldo"
        self.despesa_lancamento_path = self.fato_path / "despesa_lancamento"
        self.receita_lancamento_path = self.fato_path / "receita_lancamento"
        
        for pasta in [self.despesa_saldo_path, self.receita_saldo_path, 
                     self.despesa_lancamento_path, self.receita_lancamento_path]:
            pasta.mkdir(parents=True, exist_ok=True)
        
        print(f"📍 Base: {self.base_path.absolute()}")
        print(f"🗄️ Banco: {self.db_path}")
        print(f"📋 Metadados: {self.metadata_path}")
        
        # Arquivos de metadados
        self.mapeamento_file = self.metadata_path / "mapeamento_fatos.json"
        self.historico_file = self.metadata_path / "historico_cargas_fatos.json"
        self.calendario_file = self.metadata_path / "calendario_esperado.json"
        
        # Carregar metadados
        self.mapeamentos = self.carregar_json(self.mapeamento_file)
        self.historico = self.carregar_json(self.historico_file)
        self.calendario = self.carregar_json(self.calendario_file)
    
    def carregar_json(self, arquivo):
        """Carrega arquivo JSON se existir"""
        if arquivo.exists():
            try:
                with open(arquivo, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"⚠️ Erro ao carregar {arquivo.name}: {e}")
                return {}
        return {}
    
    def salvar_json(self, dados, arquivo):
        """Salva dados em arquivo JSON"""
        try:
            with open(arquivo, 'w', encoding='utf-8') as f:
                json.dump(dados, f, indent=2, ensure_ascii=False)
            return True
        except Exception as e:
            print(f"❌ Erro ao salvar {arquivo.name}: {e}")
            return False
    
    def calcular_hash_arquivo(self, caminho):
        """Calcula hash do arquivo para detectar mudanças"""
        hash_md5 = hashlib.md5()
        with open(caminho, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    
    def identificar_tipo_arquivo(self, nome_arquivo):
        """Identifica o tipo de tabela baseado no nome do arquivo"""
        nome_lower = nome_arquivo.lower()
        info = {
            'arquivo': nome_arquivo,
            'tipo_tabela': None,
            'tipo_carga': None
        }
        
        # Identificar tipo de tabela
        if 'despesasaldo' in nome_lower:
            info['tipo_tabela'] = 'despesa_saldo'
        elif 'receitasaldo' in nome_lower:
            info['tipo_tabela'] = 'receita_saldo'
        elif 'despesal' in nome_lower:
            info['tipo_tabela'] = 'despesa_lancamento'
        elif 'receital' in nome_lower:
            info['tipo_tabela'] = 'receita_lancamento'
        
        # Identificar se é arquivo inicial ou mensal
        if 'inicial' in nome_lower:
            info['tipo_carga'] = 'inicial'
        else:
            info['tipo_carga'] = 'mensal'
        
        return info
    
    def detectar_periodos_no_arquivo(self, caminho, tipo_tabela):
        """Detecta períodos analisando o conteúdo do arquivo"""
        try:
            # Ler amostra do arquivo
            df = pd.read_excel(caminho, nrows=1000)
            
            # Procurar colunas de exercício e mês
            colunas_upper = [col.upper() for col in df.columns]
            
            if 'COEXERCICIO' in colunas_upper and 'INMES' in colunas_upper:
                idx_exercicio = colunas_upper.index('COEXERCICIO')
                idx_mes = colunas_upper.index('INMES')
                
                df['periodo'] = (
                    df.iloc[:, idx_exercicio].astype(str) + '-' + 
                    df.iloc[:, idx_mes].astype(str).str.zfill(2)
                )
                
                periodos = sorted(df['periodo'].unique())
                return periodos
            
            # Tentar encontrar coluna PERIODO diretamente
            if 'PERIODO' in colunas_upper:
                idx_periodo = colunas_upper.index('PERIODO')
                periodos = sorted(df.iloc[:, idx_periodo].unique())
                return periodos
                
        except Exception as e:
            print(f"⚠️ Erro ao detectar períodos: {e}")
        
        return []
    
    def analisar_situacao_completa(self):
        """Análise completa da situação de todas as tabelas fato"""
        print("\n" + "="*80)
        print("🧠 ANÁLISE INTELIGENTE - TABELAS FATO")
        print("="*80)
        
        situacao = {
            'despesa_saldo': {'arquivos': [], 'total_banco': 0, 'status': {}},
            'receita_saldo': {'arquivos': [], 'total_banco': 0, 'status': {}},
            'despesa_lancamento': {'arquivos': [], 'total_banco': 0, 'status': {}},
            'receita_lancamento': {'arquivos': [], 'total_banco': 0, 'status': {}}
        }
        
        # Mapear pastas para tipos
        pastas_tipos = [
            (self.despesa_saldo_path, 'despesa_saldo'),
            (self.receita_saldo_path, 'receita_saldo'),
            (self.despesa_lancamento_path, 'despesa_lancamento'),
            (self.receita_lancamento_path, 'receita_lancamento')
        ]
        
        # Analisar arquivos em cada pasta
        for pasta, tipo in pastas_tipos:
            arquivos = list(pasta.glob("*.xlsx")) + list(pasta.glob("*.xls"))
            
            for arquivo in arquivos:
                info = self.identificar_tipo_arquivo(arquivo.name)
                info['caminho'] = arquivo
                info['hash'] = self.calcular_hash_arquivo(arquivo)
                
                # Verificar se já foi carregado
                if arquivo.name in self.mapeamentos:
                    mapa = self.mapeamentos[arquivo.name]
                    if mapa['hash'] == info['hash']:
                        info['status'] = 'sincronizado'
                    else:
                        info['status'] = 'modificado'
                else:
                    info['status'] = 'novo'
                
                situacao[tipo]['arquivos'].append(info)
        
        # Verificar totais no banco
        if self.db_path.exists():
            conn = duckdb.connect(str(self.db_path))
            try:
                for tipo in situacao.keys():
                    # Verificar se tabela existe
                    exists = conn.execute(f"""
                        SELECT COUNT(*) FROM information_schema.tables 
                        WHERE table_name = '{tipo}'
                    """).fetchone()[0] > 0
                    
                    if exists:
                        total = conn.execute(f"SELECT COUNT(*) FROM {tipo}").fetchone()[0]
                        situacao[tipo]['total_banco'] = total
            finally:
                conn.close()
        
        return situacao
    
    def gerar_calendario_esperado(self):
        """Gera calendário de arquivos esperados"""
        hoje = datetime.now()
        calendario = {}
        
        # Para cada tipo de tabela
        for tipo in ['despesa_saldo', 'receita_saldo']:
            calendario[tipo] = {
                'arquivos_esperados': [],
                'proximos_arquivos': []
            }
            
            # Arquivo inicial do ano
            calendario[tipo]['arquivos_esperados'].append({
                'nome_padrao': f'{tipo.replace("_", "")}inicial25.xlsx',
                'periodos': ['2024-01', '2024-02', '2024-03', '2024-04', '2024-05'],
                'prazo': '2025-06-01'
            })
            
            # Arquivos mensais
            for mes in range(6, 13):
                mes_nome = list(self.MESES_ABREV.keys())[mes-1]
                calendario[tipo]['arquivos_esperados'].append({
                    'nome_padrao': f'{tipo.replace("_", "")}{mes_nome}25.xlsx',
                    'periodo': f'2024-{mes:02d}',
                    'prazo': f'2025-{mes+1:02d}-05' if mes < 12 else '2026-01-05'
                })
        
        self.calendario = calendario
        self.salvar_json(calendario, self.calendario_file)
        return calendario
    
    def processar_arquivo_fato(self, info_arquivo):
        """Processa um arquivo de tabela fato"""
        arquivo = info_arquivo['arquivo']
        caminho = info_arquivo['caminho']
        tipo_tabela = info_arquivo['tipo_tabela']
        
        print(f"\n📄 Processando: {arquivo}")
        print(f"   📊 Tipo: {tipo_tabela}")
        
        try:
            # Ler arquivo completo
            print("   📖 Lendo arquivo...")
            df = pd.read_excel(caminho)
            print(f"   📊 {len(df):,} linhas, {len(df.columns)} colunas")
            
            # Padronizar nomes das colunas (lowercase)
            df.columns = df.columns.str.lower()
            
            # CORREÇÃO DE TIPOS: Converter campos numéricos para INTEGER
            # Lista de campos que devem ser INTEGER (não FLOAT)
            campos_integer = [
                'coexercicio', 'coug', 'cogestao', 'inmes', 'inesfera',
                'couo', 'cofuncao', 'cosubfuncao', 'coprograma', 
                'coprojeto', 'cosubtitulo', 'cofonte', 'conatureza',
                'comodalidade', 'coelemento', 'cogrupo', 'incategoria',
                'coevento', 'cougdestino', 'cogestaodestino',
                'cougcontab', 'cogestaocontab', 'inabreencerra',
                'nulancamento', 'intipoadm', 'cocategoriareceita',
                'cofontereceita', 'cosubfontereceita', 'corubrica',
                'coalinea', 'coclasseorc'
            ]
            
            # Aplicar conversão para campos que existem no DataFrame
            print("   🔧 Convertendo tipos de dados...")
            for campo in campos_integer:
                if campo in df.columns:
                    try:
                        # Converter para numeric primeiro (trata NaN)
                        df[campo] = pd.to_numeric(df[campo], errors='coerce')
                        # Converter para Int64 (permite NaN)
                        df[campo] = df[campo].astype('Int64')
                    except Exception as e:
                        print(f"      ⚠️ Aviso ao converter {campo}: {e}")
            
            # Campos que devem permanecer como VARCHAR/STRING
            campos_string = [
                'cocontacontabil', 'cocontacorrente', 'nudocumento',
                'incontacorrente', 'insaldocontabil', 'indebitocredito',
                'tipo_lancamento', 'hotransacao'
            ]
            
            for campo in campos_string:
                if campo in df.columns:
                    df[campo] = df[campo].astype(str).str.strip()
            
            # Campos decimais (manter como float)
            campos_decimal = ['vacredito', 'vadebito', 'valancamento',
                            'saldo_contabil_despesa', 'saldo_contabil_receita']
            
            for campo in campos_decimal:
                if campo in df.columns:
                    df[campo] = pd.to_numeric(df[campo], errors='coerce').fillna(0.0)
            
            # Adicionar timestamp de carga
            df['data_carga'] = datetime.now()
            
            # Verificar se tabela existe
            conn = duckdb.connect(str(self.db_path))
            try:
                # Verificar se tabela existe
                table_exists = conn.execute(f"""
                    SELECT COUNT(*) FROM information_schema.tables 
                    WHERE table_name = '{tipo_tabela}'
                """).fetchone()[0] > 0
                
                if table_exists:
                    # Contar registros existentes
                    count_antes = conn.execute(f"SELECT COUNT(*) FROM {tipo_tabela}").fetchone()[0]
                    print(f"   📊 Tabela existe com {count_antes:,} registros")
                    
                    print("\n   O que deseja fazer?")
                    print("   [A] ADICIONAR - Mantém os dados existentes e adiciona os novos")
                    print("   [S] SUBSTITUIR - Apaga TUDO e carrega apenas este arquivo")
                    print("   [C] CANCELAR - Não faz nada")
                    
                    resp = input("\n   Escolha (A/s/c): ").strip().lower()
                    
                    if resp == 'c':
                        print("   ⏭️ Operação cancelada")
                        return False
                    elif resp == 's':
                        # Substituir tudo
                        confirm = input("   ⚠️ CONFIRMA apagar TODOS os dados existentes? (s/N): ")
                        if confirm.lower() == 's':
                            conn.execute(f"DROP TABLE {tipo_tabela}")
                            print("   🗑️ Tabela removida")
                            table_exists = False
                        else:
                            print("   ⏭️ Operação cancelada")
                            return False
                    # Se resp == 'a' ou Enter, apenas adiciona (comportamento padrão)
                
                # Criar ou inserir na tabela
                conn.register('df_temp', df)
                
                if not table_exists:
                    print(f"   📝 Criando tabela {tipo_tabela}...")
                    conn.execute(f"CREATE TABLE {tipo_tabela} AS SELECT * FROM df_temp")
                else:
                    conn.execute(f"INSERT INTO {tipo_tabela} SELECT * FROM df_temp")
                
                conn.unregister('df_temp')
                
                # Contar registros finais
                count_final = conn.execute(f"SELECT COUNT(*) FROM {tipo_tabela}").fetchone()[0]
                print(f"   ✅ {len(df):,} registros processados!")
                print(f"   📊 Total na tabela: {count_final:,}")
                
                # Verificar tipos das colunas criadas
                print("   📋 Verificando tipos de dados...")
                tipos = conn.execute(f"""
                    SELECT column_name, data_type 
                    FROM information_schema.columns 
                    WHERE table_name = '{tipo_tabela}'
                    AND column_name IN ('couo', 'cofuncao', 'cosubfuncao', 'coprograma')
                    ORDER BY ordinal_position
                    LIMIT 4
                """).fetchall()
                
                for col, tipo in tipos:
                    print(f"      • {col}: {tipo}")
                
                # Salvar mapeamento
                self.mapeamentos[arquivo] = {
                    'tabela': tipo_tabela,
                    'tipo_carga': info_arquivo.get('tipo_carga', 'desconhecido'),
                    'hash': info_arquivo['hash'],
                    'ultima_carga': datetime.now().isoformat(),
                    'registros': len(df),
                    'colunas': list(df.columns)
                }
                self.salvar_json(self.mapeamentos, self.mapeamento_file)
                
                # Salvar histórico
                if arquivo not in self.historico:
                    self.historico[arquivo] = []
                
                self.historico[arquivo].append({
                    'data': datetime.now().isoformat(),
                    'tabela': tipo_tabela,
                    'acao': 'carga',
                    'registros': len(df)
                })
                self.salvar_json(self.historico, self.historico_file)
                
                return True
                    
            finally:
                conn.close()
                
        except Exception as e:
            print(f"   ❌ Erro: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def gerar_documentacao_estrutura(self):
        """Gera documentação da estrutura do banco"""
        print("\n📄 Gerando documentação da estrutura...")
        
        if not self.db_path.exists():
            print("❌ Banco de dados não existe ainda!")
            return False
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        estrutura_file = self.metadata_path / f"estrutura_fatos_{timestamp}.txt"
        
        conn = duckdb.connect(str(self.db_path))
        
        try:
            # Listar tabelas fato
            tabelas = conn.execute("""
                SELECT table_name 
                FROM information_schema.tables 
                WHERE table_schema = 'main'
                AND table_name IN ('despesa_saldo', 'receita_saldo', 
                                  'despesa_lancamento', 'receita_lancamento')
                ORDER BY table_name
            """).fetchall()
            
            with open(estrutura_file, 'w', encoding='utf-8') as f:
                # Cabeçalho
                f.write("=" * 100 + "\n")
                f.write("DOCUMENTAÇÃO DA ESTRUTURA - TABELAS FATO\n")
                f.write("=" * 100 + "\n\n")
                f.write(f"Banco: {self.db_path}\n")
                f.write(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                f.write(f"Total de tabelas fato: {len(tabelas)}\n\n")
                
                # Para cada tabela
                for (tabela,) in tabelas:
                    f.write("=" * 100 + "\n")
                    f.write(f"TABELA: {tabela.upper()}\n")
                    f.write("=" * 100 + "\n\n")
                    
                    # Total de registros
                    total = conn.execute(f"SELECT COUNT(*) FROM {tabela}").fetchone()[0]
                    f.write(f"Total de registros: {total:,}\n\n")
                    
                    f.write("ESTRUTURA:\n")
                    f.write("-" * 80 + "\n")
                    
                    # Estrutura da tabela
                    colunas = conn.execute(f"""
                        SELECT column_name, data_type, is_nullable
                        FROM information_schema.columns
                        WHERE table_name = '{tabela}'
                        ORDER BY ordinal_position
                    """).fetchall()
                    
                    f.write(f"{'#':<5} {'Campo':<30} {'Tipo':<20} {'Nulo?':<10}\n")
                    f.write("-" * 65 + "\n")
                    
                    for i, (col, tipo, nullable) in enumerate(colunas, 1):
                        nullable_str = "SIM" if nullable == "YES" else "NÃO"
                        f.write(f"{i:<5} {col[:30]:<30} {tipo[:20]:<20} {nullable_str:<10}\n")
                    
                    f.write("\n")
            
            print(f"✅ Documentação gerada: {estrutura_file}")
            return True
            
        except Exception as e:
            print(f"❌ Erro: {e}")
            return False
        finally:
            conn.close()
    
    def gerar_amostra_excel(self, tabela, qtd_registros=1000):
        """Gera arquivo Excel com amostra de dados"""
        if not self.db_path.exists():
            print("❌ Banco não existe!")
            return False
        
        conn = duckdb.connect(str(self.db_path))
        
        try:
            # Verificar se tabela existe
            exists = conn.execute(f"""
                SELECT COUNT(*) FROM information_schema.tables 
                WHERE table_name = '{tabela}'
            """).fetchone()[0] > 0
            
            if not exists:
                print(f"❌ Tabela {tabela} não existe!")
                return False
            
            # Buscar total e amostra
            total = conn.execute(f"SELECT COUNT(*) FROM {tabela}").fetchone()[0]
            
            if total == 0:
                print(f"⚠️ Tabela {tabela} está vazia!")
                return False
            
            qtd_buscar = min(qtd_registros, total)
            
            print(f"📊 Gerando amostra de {qtd_buscar:,} registros de {tabela}...")
            
            # Buscar dados
            df = pd.read_sql(f"""
                SELECT * FROM {tabela} 
                LIMIT {qtd_buscar}
            """, conn)
            
            # Gerar arquivo Excel
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            arquivo_excel = self.amostras_path / f"amostra_{tabela}_{timestamp}.xlsx"
            
            with pd.ExcelWriter(arquivo_excel, engine='openpyxl') as writer:
                # Aba com dados
                df.to_excel(writer, sheet_name='Dados', index=False)
                
                # Aba com resumo
                resumo_data = {
                    'Informação': [
                        'Tabela',
                        'Total de registros no banco',
                        'Registros nesta amostra',
                        'Data da extração'
                    ],
                    'Valor': [
                        tabela,
                        f"{total:,}",
                        f"{qtd_buscar:,}",
                        datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                    ]
                }
                
                df_resumo = pd.DataFrame(resumo_data)
                df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
                
                # Formatar
                workbook = writer.book
                
                # Formatar aba Resumo
                ws_resumo = workbook['Resumo']
                for cell in ws_resumo['A1:B1']:
                    for c in cell:
                        c.font = Font(bold=True)
                        c.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        c.font = Font(color="FFFFFF", bold=True)
                
                # Formatar aba Dados
                ws_dados = workbook['Dados']
                for cell in ws_dados[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
            
            print(f"✅ Amostra gerada: {arquivo_excel}")
            print(f"   • {qtd_buscar:,} registros exportados")
            
            return True
            
        except Exception as e:
            print(f"❌ Erro: {e}")
            return False
        finally:
            conn.close()
    
    def menu_principal(self):
        """Menu principal do sistema"""
        while True:
            # Análise da situação
            situacao = self.analisar_situacao_completa()
            
            print("\n" + "="*80)
            print("🧠 CARREGADOR INTELIGENTE - TABELAS FATO")
            print("="*80)
            
            # Resumo por tabela
            for tipo, info in situacao.items():
                arquivos = info['arquivos']
                total_banco = info['total_banco']
                
                print(f"\n📊 {tipo.upper().replace('_', ' ')}:")
                
                if total_banco > 0:
                    print(f"   ✅ Total no banco: {total_banco:,} registros")
                else:
                    print(f"   ⚠️ Tabela vazia ou não existe")
                
                novos = [a for a in arquivos if a.get('status') == 'novo']
                modificados = [a for a in arquivos if a.get('status') == 'modificado']
                
                if novos:
                    print(f"   🆕 Arquivos novos: {len(novos)}")
                if modificados:
                    print(f"   📝 Arquivos modificados: {len(modificados)}")
            
            # Menu
            print("\n" + "="*80)
            print("OPÇÕES:")
            print("[1] Carregar arquivos novos")
            print("[2] Atualizar arquivos modificados")
            print("[3] Processar arquivo específico")
            print("[4] Gerar amostra Excel (1000 registros)")
            print("[5] Gerar documentação da estrutura")
            print("[6] Ver histórico de cargas")
            print("[7] Ver calendário esperado")
            print("[8] Análise de integridade")
            print("[0] Sair")
            
            opcao = input("\n➤ Escolha: ").strip()
            
            if opcao == '0':
                break
            
            elif opcao == '1':
                # Carregar novos
                for tipo, info in situacao.items():
                    novos = [a for a in info['arquivos'] if a.get('status') == 'novo']
                    for arquivo_info in novos:
                        print(f"\n🆕 Novo arquivo detectado: {arquivo_info['arquivo']}")
                        print(f"   Tipo: {tipo}")
                        resp = input("   Carregar? (S/n): ")
                        if resp.lower() != 'n':
                            self.processar_arquivo_fato(arquivo_info)
            
            elif opcao == '2':
                # Atualizar modificados
                for tipo, info in situacao.items():
                    modificados = [a for a in info['arquivos'] if a.get('status') == 'modificado']
                    for arquivo_info in modificados:
                        print(f"\n📝 Arquivo modificado: {arquivo_info['arquivo']}")
                        resp = input("   Atualizar? (s/N): ")
                        if resp.lower() == 's':
                            self.processar_arquivo_fato(arquivo_info)
            
            elif opcao == '3':
                # Processar específico
                print("\nArquivos disponíveis:")
                todos_arquivos = []
                for tipo, info in situacao.items():
                    for arq in info['arquivos']:
                        todos_arquivos.append(arq)
                        print(f"  [{len(todos_arquivos)}] {arq['arquivo']} ({tipo})")
                
                if todos_arquivos:
                    try:
                        idx = int(input("\nEscolha o número: ")) - 1
                        if 0 <= idx < len(todos_arquivos):
                            self.processar_arquivo_fato(todos_arquivos[idx])
                    except:
                        print("❌ Opção inválida!")
            
            elif opcao == '4':
                # Gerar amostra
                print("\nTabelas disponíveis:")
                tabelas = ['despesa_saldo', 'receita_saldo', 'despesa_lancamento', 'receita_lancamento']
                for i, t in enumerate(tabelas, 1):
                    print(f"  [{i}] {t}")
                
                try:
                    idx = int(input("\nEscolha: ")) - 1
                    if 0 <= idx < len(tabelas):
                        qtd = input("Quantos registros? (padrão 1000): ").strip()
                        qtd = int(qtd) if qtd else 1000
                        self.gerar_amostra_excel(tabelas[idx], qtd)
                except:
                    print("❌ Opção inválida!")
            
            elif opcao == '5':
                self.gerar_documentacao_estrutura()
            
            elif opcao == '6':
                # Ver histórico
                print("\n📜 HISTÓRICO DE CARGAS:")
                for arquivo, eventos in self.historico.items():
                    print(f"\n{arquivo}:")
                    for evento in eventos[-5:]:  # Últimos 5 eventos
                        print(f"  • {evento['data'][:19]} - {evento['tabela']} - "
                              f"{evento['registros']:,} registros")
            
            elif opcao == '7':
                # Calendário esperado
                print("\n📅 ARQUIVOS ESPERADOS (padrão de nomenclatura):")
                print("\nDESPESA SALDO:")
                print("  • despesasaldoinicial25.xlsx - Dados iniciais do ano")
                print("  • despesasaldojun25.xlsx - Junho")
                print("  • despesasaldojul25.xlsx - Julho")
                print("  • despesasaldoago25.xlsx - Agosto")
                print("  • ... e assim por diante")
                print("\nRECEITA SALDO:")
                print("  • receitasaldoinicial25.xlsx - Dados iniciais do ano")
                print("  • receitasaldojun25.xlsx - Junho")
                print("  • ... mesmo padrão da despesa")
            
            elif opcao == '8':
                # Análise de integridade
                print("\n🔍 ANÁLISE DE INTEGRIDADE:")
                
                if self.db_path.exists():
                    conn = duckdb.connect(str(self.db_path))
                    try:
                        for tipo in ['despesa_saldo', 'receita_saldo', 'despesa_lancamento', 'receita_lancamento']:
                            exists = conn.execute(f"""
                                SELECT COUNT(*) FROM information_schema.tables 
                                WHERE table_name = '{tipo}'
                            """).fetchone()[0] > 0
                            
                            if exists:
                                print(f"\n{tipo.upper()}:")
                                
                                # Total de registros
                                total = conn.execute(f"SELECT COUNT(*) FROM {tipo}").fetchone()[0]
                                print(f"  ✅ Total de registros: {total:,}")
                                
                                # Verificar colunas principais
                                colunas = conn.execute(f"""
                                    SELECT COUNT(*) 
                                    FROM information_schema.columns
                                    WHERE table_name = '{tipo}'
                                """).fetchone()[0]
                                print(f"  ✅ Total de colunas: {colunas}")
                                
                                # Verificar valores nulos em colunas chave (se existirem)
                                colunas_chave = ['coexercicio', 'coug', 'inmes']
                                for col in colunas_chave:
                                    try:
                                        nulos = conn.execute(f"""
                                            SELECT COUNT(*) 
                                            FROM {tipo} 
                                            WHERE {col} IS NULL
                                        """).fetchone()[0]
                                        if nulos > 0:
                                            print(f"  ⚠️ Valores nulos em {col}: {nulos}")
                                    except:
                                        pass  # Coluna pode não existir
                    finally:
                        conn.close()
                else:
                    print("❌ Banco de dados não existe!")
            
            else:
                print("❌ Opção inválida!")
        
        print("\n✨ Sistema encerrado!")

def main():
    """Função principal"""
    parser = argparse.ArgumentParser(description='Carregador Inteligente de Tabelas Fato')
    parser.add_argument('--estrutura', action='store_true', 
                       help='Apenas gera documentação da estrutura')
    parser.add_argument('--amostra', type=str, 
                       help='Gera amostra Excel da tabela especificada')
    
    args = parser.parse_args()
    
    carregador = CarregadorInteligenteTabelas()
    
    if args.estrutura:
        carregador.gerar_documentacao_estrutura()
    elif args.amostra:
        carregador.gerar_amostra_excel(args.amostra)
    else:
        carregador.menu_principal()

if __name__ == "__main__":
    main()